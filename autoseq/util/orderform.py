import logging

from openpyxl import load_workbook


def parse_orderform_block(block_of_values):
    """
    Extract clinseq barcodes from the given list of order form fields. Looks
    in the entries between <SAMPLE ENTRIES> and </SAMPLE ENTRIES> for clinseq barcodes.

    :param block_of_values: List of fields from which to extract clinseq barcodes
    :return: List of (not-yet validated) clinseq barcode strings
    """

    clinseq_barcode_strings = []
    clinseq_barcodes_section = False
    for cell_value in block_of_values:
        if cell_value == "</SAMPLE ENTRIES>":
            clinseq_barcodes_section = False
        if clinseq_barcodes_section:
            clinseq_barcode_strings.append(cell_value)
        if cell_value == "<SAMPLE ENTRIES>":
            clinseq_barcodes_section = True
        else:
            logging.debug("Ignoring field from order form: " + cell_value)

    return clinseq_barcode_strings


def parse_orderform_worksheet(order_form_worksheet):
    """
    Extract clinseq barcodes from the given order form excel spreadsheet worksheet.

    :param order_form_worksheet: An openpyxl worksheet.
    :return: List of clinseq barcodes extracted from the worksheet.
    """

    first_column_vals = [row[0].value for row in list(order_form_worksheet.iter_rows()) if
                         row[0].value is not None]

    return parse_orderform_block(first_column_vals)


def parse_orderform(order_form_filename):
    """
    Extract clinseq barcodes from the specified order form.

    :param order_form_filename: An excel spreadsheet filename.
    :return: List of clinseq barcodes extracted from the order form.
    """

    workbook = load_workbook(order_form_filename)
    return parse_orderform_worksheet(workbook.worksheets[0])
